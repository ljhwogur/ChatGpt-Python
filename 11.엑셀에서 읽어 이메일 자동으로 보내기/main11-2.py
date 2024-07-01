import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl

# 엑셀 파일 경로
file_path = "11.엑셀에서 읽어 이메일 자동으로 보내기/이메일.xlsx"

# 이메일 계정 정보
smtp_server = "smtp.naver.com"
smtp_port = 587
email_address = "ljhwogur1@naver.com"
email_password = "83Wogur."

# 엑셀 파일 열기
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# A열과 B열의 값 읽기
emails_and_names = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
    email, name = row
    emails_and_names.append((email, name))
    
# 이메일 보내기 함수
def send_email(to_email, name):
    subject = f"{name}님 환영합니다."
    body = f"{name}님 늦지 않게 와주세요."
    
    msg = MIMEMultipart()
    msg['From'] = email_address
    msg['To'] = to_email
    msg['Subject'] = subject
    
    msg.attach(MIMEText(body, 'plain'))
    
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(email_address, email_password)
        text = msg.as_string()
        server.sendmail(email_address, to_email, text)
        server.quit()
        print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"Failed to send email to {to_email}: {e}")

# 모든 이메일 주소로 이메일 보내기
for email, name in emails_and_names:
    send_email(email, name)