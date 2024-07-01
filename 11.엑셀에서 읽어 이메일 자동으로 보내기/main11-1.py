import openpyxl

# 엑셀 파일 경로
file_path = "11.엑셀에서 읽어 이메일 자동으로 보내기/이메일.xlsx"

# 엑셀 파일 열기
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# A열과 B열의 값 읽기
emails_and_names = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
    email, name = row
    emails_and_names.append((email, name))
    
# 결과 출력
print("Emails and Names:")
for email, name in emails_and_names:
    print(f"Email: {email}, Name: {name}")